from openpyxl import load_workbook

# 打开文件
wb = load_workbook("山东高速集团校园招聘岗位2022.xlsx")

print(wb.sheetnames)  # 获取所有的sheet名

sheet = wb.get_sheet_by_name("岗位计划表")  # 获取第一张表的对象，可以对其进行操作

# # 获取数据
# print(sheet["B5"])
# print(sheet["B5"].value)  # 获取B5的值

# # 获取指定列的切片数据
# print(sheet["B5:B10"])  # ((B5,),(B6,),(B7,),(B8,),(B9,),(B10,))先行后列，先取出来的是行
# for cell in sheet["B5:B10"]:  # cell是一个元组
#     print(cell[0].value)

# # 按行遍历
# for row in sheet:  # 循环获取表数据
#     for cell in row:  # 循环获取每个单元格数据
#         print(cell.value, end=",")
#     print()
#
# # 按列遍历：A1，A2，A3这样的顺序
# for column in sheet.columns:
#     for cell in column:
#         print(cell.value, end=",")
#     print()
#
# 遍历指定行&列
for row in sheet.iter_rows(min_row=5, max_row=10, max_col=5):  # 从第2行到第5行，每行打印5列
    for cell in row:
        print(cell.value, end=",")
    print()

# 遍历指定列
# for col in sheet.iter_cols(min_col=2, max_col=5, ):  # 从第2列到第5列
#     for i in col:
#         print(i.value, end=",")
#     print()


# # 删除工作表
# wb.remove(sheet)

# # 保存文件
# wb.save("文件名.xlsx")
